# -*- encoding: utf-8 -*-
#!/usr/bin/python3

import os
import sys
import csv
import string
import re
from turtle import clear

import xlrd
import xlwt
import openpyxl
import codecs
import pandas as pd
from openpyxl.utils import get_column_letter
from createVbo import create_vbo

# 读取文件并打印到屏幕
def readfile(filename):
    '''Print a file to the standard output.'''
    f = open(filename, 'r', encoding="utf-8")
    while True:
        line = f.readline()
        if len(line) == 0:
            break
        print(line)
        print("complete")
    f.close()

# 删除laptime table

def delet_laptime_table(filename):
    f = open(filename, 'r')
    new_file = filename + ".nolap.xls"
    fw = open(new_file, 'w')
    laptime = "Lap:"
    line_table = 0
    lines = f.readlines()
    for line_table in lines:
        if laptime in line_table:
            if re.match("^Lap:$", line_table, flags=0) != None:
                break
        fw.write(line_table)
    fw.close()
    f.close()
    print("new xls file is: ", new_file)
    return new_file

# def txt_xls(filename):
#     try:
#         f = open(filename)
#         xls = xlwt.Workbook()
#         #⽣成excel的⽅法，声明excel
#         sheet = xls.add_sheet('Sheet', cell_overwrite_ok=True)
#         x = 0
#         print(filename, " is read")
#         while True:
#             #按⾏循环，读取⽂本⽂件
#             line = f.readline()
#             if not line:
#                 break  # 如果没有内容，则退出循环
#             for i in range(len(line.split('\t'))):
#                 item = line.split('\t')[i]
#                 sheet.write(x, i, item)  # x单元格经度，i 单元格纬度
#             x += 1  # excel另起⼀⾏
#         f.close()
#         new_name = filename + ".new.xls"
#         xls.save(new_name)  # 保存xls⽂件
#         print(new_name, " is saved")
#         return new_name
#     except:
#         raise

# def xlsx_to_txt(filename):
#     df = pd.read_excel(filename, sheetname='Sheet', header=None)  # 使⽤pandas模块读取数据
#     print('开始写⼊txt⽂件...')
#     df.to_csv(filename + ".txt", header=None, sep=' ', index=False)  # 写⼊，逗号分隔
#     print('⽂件写⼊成功!')


def txt_to_xlsx(filename):
    fr = codecs.open(filename, 'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    row = 0
    for line in fr:
        row += 1
        # line = line.strip()
        line = line.split('\t')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(
                get_column_letter(col)))
    outfile = filename + ".new.xlsx"
    wb.save(outfile)
    print(outfile, " is saved")
    return outfile

# 删除aim导出的xls的前两列， 并替换Time (s)为time
def del_first_2_col(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.delete_cols(idx=1, amount=2)
    ws.cell(column=1, row=1).value = "time"
    outfile = filename + ".new.xlsx"
    wb.save(outfile)
    print(outfile, " is saved")
    return outfile

    
def convert_2_vbo(txt_file, xlsx_file):
    # 将第一行的表头写进vbo的表头[header]
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    file = open(txt_file, 'a')
    for cell in ws[1]:
        print(cell.value)
        file.write(cell.value)
        file.write("\n")
    
    # [comments] 写入
    file.write("[comments]\n")
    file.write("Script generated by Mok\n")
    file.write("Script version: 1.0.0\n")
    
    # 写入 [column names]
    file.write("[column names]\n")
    for cell in ws[1]:
        print(cell.value)
        file.write(cell.value)
        file.write(" ")
    file.write("\n")
    
    # 写入 [data]
    # 从第二行开始读取
    file.write("[data]\n")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            print(cell.value)
            file.write(cell.value)
            file.write(" ")
        file.write("\n")
    file.close()
    return txt_file


if __name__ == '__main__':
    # 命令行第二个参数定义输入文件名路径
    print("This is a data file exported form Aim \n")
    print("read file:", sys.argv[1])
    print(os.path.dirname(os.path.abspath(sys.argv[1])))
    new = delet_laptime_table(sys.argv[1])

    newrow = txt_to_xlsx(new)
    os.remove(new)
    final_xlsx = del_first_2_col(newrow)
    os.remove(newrow)
    # os.rename(final_xlsx, sys.argv[1] + ".final.xlsx")
    # print("final xlsx is: ", sys.argv[1] + ".final.xlsx")
    print("Creating Vbo file ...")
    new_vbo_name = create_vbo(sys.argv[1])
    final_vbo = convert_2_vbo(new_vbo_name, final_xlsx)
    os.remove(final_xlsx)
    os.system("cls")
    print("Successfully converted to Vbo file\n")
    print("The vbo file is in: ", final_vbo)
    

    

