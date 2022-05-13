# -*- encoding: utf-8 -*-
#!/usr/bin/python3

import os
import sys

import csv
import string

import xlrd
import xlwt
import openpyxl
import codecs
from openpyxl.utils import get_column_letter

import re
import time


# 读取文件并打印到屏幕
def readfile(filename):
    '''Print a file to the standard output.'''
    f = open(filename, 'r', encoding="utf-8")
    while True:
        line = f.readline()
        if len(line) == 0:
            break
        print(line)
        print("complete")
    f.close()

# 删除laptime table

def delet_laptime_table(filename):
    f = open(filename, 'r')
    new_file = filename + ".nolap.xls"
    fw = open(new_file, 'w')
    laptime = "Lap:"
    line_table = 0
    lines = f.readlines()
    for line_table in lines:
        if laptime in line_table:
            if re.match("^Lap:$", line_table, flags=0) != None:
                break
        fw.write(line_table)
    fw.close()
    f.close()
    print("new xls file is: ", new_file)
    return new_file

# 创建vbo文件
def creat_vbo(filename):
    full_path = filename + ".vbo"
    file = open(full_path, 'w', encoding="utf-8")
    new_str = "File created on" + time.time()
    file.write(new_str)
    file.close()

# def txt_xls(filename):
#     try:
#         f = open(filename)
#         xls = xlwt.Workbook()
#         #⽣成excel的⽅法，声明excel
#         sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
#         x = 0
#         print(filename, " is read")
#         while True:
#             #按⾏循环，读取⽂本⽂件
#             line = f.readline()
#             if not line:
#                 break  # 如果没有内容，则退出循环
#             for i in range(len(line.split('\t'))):
#                 item = line.split('\t')[i]
#                 sheet.write(x, i, item)  # x单元格经度，i 单元格纬度
#             x += 1  # excel另起⼀⾏
#         f.close()
#         new_name = filename + ".new.xls"
#         xls.save(new_name)  # 保存xls⽂件
#         print(new_name, " is saved")
#         return new_name
#     except:
#         raise


def txt_to_xlsx(filename):
    fr = codecs.open(filename, 'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    # ws = wb.create_sheet()
    # ws.title = 'Sheet'
    row = 0
    for line in fr:
        row += 1
        # line = line.strip()
        line = line.split('\t')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(
                get_column_letter(col)))
    outfile = filename + ".new.xlsx"
    wb.save(outfile)
    print(outfile, " is saved")
    return outfile

# 删除aim导出的xls的前两列
def del_first_2_col(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.delete_cols(idx=1, amount=2)
    outfile = filename + ".new.xlsx"
    wb.save(outfile)
    print(outfile, " is saved")
    return outfile


if __name__ == '__main__':
    print("This is a data file exported form Aim \n")
    print("read file:", sys.argv[1])
    print(os.path.dirname(os.path.abspath(sys.argv[1])))
    new = delet_laptime_table(sys.argv[1])

    newrow = txt_to_xlsx(new)
    os.remove(new)
    final_xlsx = del_first_2_col(newrow)
    os.remove(newrow)
    # os.rename(final_xlsx, sys.argv[1] + ".final.xlsx")
    # print("final xlsx is: ", sys.argv[1] + ".final.xlsx")
    


