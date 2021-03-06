# -*- encoding: utf-8 -*-
#!/usr/bin/python3

import openpyxl
import codecs
from openpyxl.utils import get_column_letter

def txt_to_xlsx(filename, outfile):
    fr = codecs.open(filename, 'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws = wb.create_sheet()
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('\t')
        col = 0
        for j in range(len(line)):
            col += 1
            #print (line[j])
            ws.cell(column=col, row=row, value=line[j].format(
                get_column_letter(col)))
    wb.save(outfile)
#读取xlsx内容

def read_xlsx(filename):
    #载⼊⽂件
    wb = openpyxl.load_workbook(filename)
    #获取Sheet1⼯作表
    ws = wb.get_sheet_by_name('Sheet')
    #按⾏读取
    for row in ws.rows:
        for cell in row:
            print(cell.value)
    #按列读
    for col in ws.columns:
        for cell in col:
            print(cell.value)


if __name__ == '__main__':
    inputfileTxt = 'test.txt'
    outfileExcel = 'text_result.xlsx'
    txt_to_xlsx(inputfileTxt, outfileExcel)
    read_xlsx(outfileExcel)

